from PIL import Image, ImageDraw, ImageFont
import random
from openpyxl import load_workbook
import os
from threading import Thread
from time import time


def get_coors(co_wb_loc, sheet_name):
    coors_wb = load_workbook(co_wb_loc)
    st = coors_wb[sheet_name]
    coors = []
    temp = []
    for i in range(2, st.max_row+1):
        temp.append(st.cell(row=i, column=2).value)
        temp.append(st.cell(row=i, column=3).value)
        temp.append(st.cell(row=i, column=1).value)
        coors.append(temp)
        temp = []
    return coors


def comb_pdf(loc):
    file_list = os.listdir(loc+'.')
    pic_name = []
    im_list = []
    for x in file_list:
        if "jpg" in x or 'png' in x or 'jpeg' in x:
            pic_name.append(x)
            
    for i in pic_name:
        img = Image.open(loc+i)
        if img.mode == "RGBA":
            img = img.convert('RGB')
            img.save(loc+i)
            
    new_pic = []
    for x in pic_name:
        if "jpg" in x:
            new_pic.append(loc+x)
    for x in pic_name:
        if "png" in x:
            new_pic.append(loc+x)
    im1 = Image.open(new_pic[0])
    new_pic.pop(0)
    for i in new_pic:
        img = Image.open(i)
        im_list.append(img)
    im1.save(loc+'Doc.pdf', "PDF", resolution=100.0, save_all=True, append_images=im_list)


def read_from_excel(key_word, wb, x):
    st = wb['Sheet1']
    try:
        for i in range(1, st.max_column+1):
            if st.cell(1, i).value == key_word:
                text = st.cell(x, i).value
        return text
    except:
        return ' '


def add_info_to_pic(img, coors, pic_save_to, x, order):
    image = Image.open(img)
    wb = load_workbook(r'N-db.xlsx')

    font = ImageFont.truetype('C:/windows/fonts/Arial.ttf', 22)
    fillcolor = "#000000"

    input_text = []
    ans_title = []

    for i in range(0, len(coors)):
        try:
            info = read_from_excel(coors[i][2], wb, x)
            input_text.append(info)
            ans_title.append(coors[i][2])
            draw = ImageDraw.Draw(image)
            draw.text((int(coors[i][0]), int(coors[i][1])), info, font=font, fill=fillcolor)
        except:
            pass

    image.save(pic_save_to+"\\"+str(order)+'.png', 'png')
    return input_text, ans_title


def create_doc(template_loc, coors, save_to, order):
    x = random.randint(2, 1000)
    ans = add_info_to_pic(template_loc, coors, save_to, x, order)
    
    ans_info = ans[0]
    ans_title = ans[1]
    ans = []
    for i in range(0, len(ans_info)):
        ans.append(str(ans_title[i])+'~'+str(ans_info[i]))
    f = open(save_to+'\\'+'Ans'+'.txt', 'a')
    f.write("\n".join(ans))


def start(save_to):
    co_wb_loc = r'E:\Python\Training\Project - EB COE - Enrollment\Training Material info.xlsx'

    main_wb = load_workbook(co_wb_loc)
    st = main_wb[r'Sheet1']

    same_carrier = []
    whole_loop = []
    temp = []

    i = 2
    while i < st.max_row+1:
        o = 0
        while o < st.max_row+1:
            if st.cell(i+o, 2).value == st.cell(i, 2).value:
                temp.append(st.cell(i+o, 1).value)
                temp.append(st.cell(i+o, 2).value)
                temp.append(st.cell(i+o, 3).value)
                temp.append(st.cell(i+o, 4).value)
                temp.append(st.cell(i+o, 5).value)
                temp.append(st.cell(i+o, 6).value)
                temp.append(st.cell(i+o, 7).value)
                if temp:
                    same_carrier.append(temp)
                    temp = []
            elif st.cell(i+o, 2).value != st.cell(i, 2).value:
                x = o
                break
            o = o+1

        whole_loop.append(same_carrier)
        same_carrier = []

        i += x
        if i == st.max_row:
            if st.cell(st.max_row, 2).value != st.cell(st.max_row - 1, 2).value:
                temp.append(st.cell(st.max_row, 1).value)
                temp.append(st.cell(st.max_row, 2).value)
                temp.append(st.cell(st.max_row, 3).value)
                temp.append(st.cell(st.max_row, 4).value)
                temp.append(st.cell(st.max_row, 5).value)
                temp.append(st.cell(st.max_row, 6).value)
                temp.append(st.cell(st.max_row, 7).value)
                if temp:
                    same_carrier.append(temp)
                    whole_loop.append(same_carrier)
            break

    for carrier in whole_loop:
        carrier_name = carrier[0][1]
        os.mkdir(save_to + '\\' + str(carrier_name))

        for tp in carrier:
            tp_id = tp[0]
            tp_lc = tp[2]
            tp_name = tp[3]
            coor_loc = tp[5]
            coor_st_name = tp[6]
            coors = get_coors(coor_loc, coor_st_name)
            create_doc(tp_lc + '\\' + tp_name, coors, save_to + '\\' + str(carrier_name), tp_id)
        comb_pdf(save_to + '\\' + str(carrier_name) + '\\')


def main():
    process_list = []
    i = 1
    while i < 6:
        save_main = r'E:\Python\Training\Project - EB COE - Enrollment\temp'+'\\'+str(i)
        os.mkdir(save_main)
        process_list.append(save_main)
        i += 1

    ts = []
    for item in process_list:
        t = Thread(target=start, args=[item])
        t.start()
        ts.append(t)
    for t in ts:
        t.join()


if __name__ == '__main__':
    start_time = time()
    main()
    end = time()
    print('Cost {} seconds'.format((end - start_time)))