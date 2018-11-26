from PIL import Image, ImageDraw, ImageFont
from pylab import *
from openpyxl import load_workbook
import shutil, os

cor = 'E:\\x.png'
wb_loc = r'E:\Python\Training\Project - EB COE - Enrollment\Coors.xlsx'

cor_backup = os.path.dirname(cor) + 'with marks - ' + os.path.basename(cor)
shutil.copyfile(cor, cor_backup)


def get_cor(img_loc):
    im = array(Image.open(img_loc))
    imshow(im)

    x = []
    y = []
    s = []

    try:
        while True:
            aloc = ginput()
            s.append([round(aloc[0][0], 2), round(aloc[0][1], 2)])
            x.append(round(aloc[0][0], 2))
            y.append(round(aloc[0][1], 2))
    except:
        return s


cos = get_cor(cor_backup)

font = ImageFont.truetype('C:/windows/fonts/Arial.ttf', 22)
fillcolor = "#000000"
image = Image.open(cor_backup)
for c in cos:
    draw = ImageDraw.Draw(image)
    draw.text((int(c[0]), int(c[1])), str(cos.index(c) + 1), font=font, fill=fillcolor)
image.save(cor_backup, 'png')

coors_wb = load_workbook(wb_loc)
st = coors_wb.create_sheet()
st.cell(1, 1).value = os.path.basename(cor)[:re.search('\.', os.path.basename(cor)).span()[0]]

for i in range(0, len(cos)):
    st.cell(i + 2, 1).value = str(i + 1)
    st.cell(i + 2, 2).value = cos[i][0]
    st.cell(i + 2, 3).value = cos[i][1]


coors_wb.save(wb_loc)
